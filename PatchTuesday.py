import requests
import sys
import argparse
import time
from datetime import datetime, timedelta
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font
from bs4 import BeautifulSoup
import re

FEEDLY_URL = "https://feedly.com/cve/{}"

MSRC_API_CVRF_URL = "https://api.msrc.microsoft.com/cvrf/v3.0/cvrf/{}"
MSRC_API_CVE_URL = "https://msrc.microsoft.com/update-guide/en-US/vulnerability/{}"
HEADERS = {"Accept": "application/json"}

ADOBE_SECURITY_BULLETIN_URL = "https://helpx.adobe.com/security/security-bulletin.html"
# DO NOT USE AU Adobe URL. They don't update it promptly, I guess Adobe hates Australians
ADOBE_BASE_URL = "https://helpx.adobe.com"

SAP_SECURITY_NOTES_URL = "https://support.sap.com/en/my-support/knowledge-base/security-notes-news/{}.html"

ORACLE_SECURITY_NOTES_URL = "https://www.oracle.com/security-alerts/cpu{}.html"

# General reusable functions across Vendors
# - save_to_excel - given the rows and filename, save the file as xlsx
# - get_patch_tuesday - given the users input, get the date of patch tuesday that month
# - make_vendor_request_JSON - just makes the API request and returns the response in JSON
# - make_vendor_request_HTML - for web scraping, just grab the website content as HTML
# - check_feedly - given a CVE, get the POC/Exploited status from Feedly
# - enrich_cves_with_feedly - check a list of CVEs with Feedly 


def save_to_excel(rows, filename):
    if not rows:
        print(f"[!] No content found for {filename}. No file written.")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Vulnerabilities"

    headers = ["CVE", "Title", "CVSS Score", "Public PoC", "Exploited", "Affected Products", "Publish Date", "Release Notes"] # RELEASE NOTE MUST ALWAYS BE LAST COLUMN
    ws.append(headers)

    for row_data in rows:
        ws.append([row_data.get(h, "") for h in headers])

        current_row = ws.max_row
        url_cell = ws.cell(row=current_row, column=len(headers))  # last column (Release Note)

        url = row_data.get("Release Notes")
        if url:
            url_cell.value = "Release Note"          # display text
            url_cell.hyperlink = url                  # actual hyperlink
            url_cell.font = Font(underline="single", color="0000FF")

    try:
        wb.save(filename)
    except PermissionError:
        print(f"[!] Permission denied while saving '{filename}'. Is it open in Excel?")
        return
    except Exception as e:
        print(f"[!] Failed to save Excel file '{filename}': {e}")
        return
    print(f"[+] Done. Output saved to '{filename}'.")

def get_patch_tuesday(target_year, target_month):
    # Find the first day of the month
    first_day = datetime(target_year, target_month, 1)
    # Get the weekday of the first day (0=Monday, 1=Tuesday, ...)
    first_weekday = first_day.weekday()

    # Calculate how many days until the first Tuesday
    days_until_tuesday = (1 - first_weekday) % 7
    first_tuesday = first_day + timedelta(days=days_until_tuesday)

    # Add 7 days to get the second Tuesday (Patch Tuesday)
    patch_tuesday = first_tuesday + timedelta(days=7)
    return patch_tuesday.date()

def make_vendor_request(URL, month_selected, type):
    try:
        if type == "HTML":
            response = requests.get(URL)
            size_bytes = len(response.content)
            print(f"[*] Retrieved '{size_bytes}' bytes of data")
            response.raise_for_status()  # raises HTTPError for bad status codes
            return response
        elif type == "JSON":
            response = requests.get(URL, headers=HEADERS)
            size_bytes = len(response.content)
            print(f"[*] Retrieved '{size_bytes}' bytes of data")
            response.raise_for_status()  # raises HTTPError for bad status codes
            return response.json()
    except requests.exceptions.HTTPError as http_err:
        if response.status_code == 404:
            print(f"[404] Patch Tuesday release for month '{month_selected}' not found.")
        else:
            print(f"[HTTP {response.status_code}] Error fetching release '{month_selected}': {http_err}")
        return
    except requests.exceptions.RequestException as e:
        print(f"[!] Connection error or invalid response: {e}")
        return

def check_feedly(cve):
    print(f"[*] Fetching Feedly enrichment for {cve}...")
    response = make_vendor_request(FEEDLY_URL.format(cve), "FEEDLY ERROR", "HTML")
    soup = BeautifulSoup(response.text, "html.parser")
    if not soup or not soup.body:
        print(f"[!] Empty or malformed Feedly response for {cve}")
        return {"CVE": cve, "Public PoC": "Unknown", "Exploited": "Unknown"}

    # Look for <h3> containing "exploitation" (case-insensitive)
    exploitation_header = soup.find("h3", string=lambda s: s and "exploitation" in s.lower())

    no_poc_str = "There is no evidence that a public proof-of-concept exists."
    no_exploit_str = "There is no evidence of proof of exploitation at the moment."

    if not exploitation_header:
        # Exploitation section missing → default values
        public_poc = "No"
        exploited = "No"
    else:
        # Get the text content of the exploitation section (next sibling elements)
        section_text = ""
        next_el = exploitation_header.find_next()
        while next_el and next_el.name != "h3":
            if isinstance(next_el, str):
                section_text += next_el.strip() + " "
            elif next_el.name in ["p", "div", "span"]:
                section_text += next_el.get_text(separator=" ", strip=True) + " "
            next_el = next_el.find_next()

        section_text = section_text.lower()

        if not section_text.strip():
            print(f"[!] Exploitation section exists but is empty for {cve}")
            public_poc = "Unknown"
            exploited = "Unknown"
        else:
            public_poc = "No" if no_poc_str.lower() in section_text else "Yes"
            exploited = "No" if no_exploit_str.lower() in section_text else "Yes"

    return {
        "CVE": cve,
        "Public PoC": public_poc,
        "Exploited": exploited
    }

def enrich_cves_with_feedly(cve_list):
    for cve_entry in cve_list:
        cve_id = cve_entry.get("CVE")
        if not cve_id:
            continue

        try:
            feedly_info = check_feedly(cve_id)
            # Add to the existing CVE record
            cve_entry["Public PoC"] = feedly_info["Public PoC"]
            cve_entry["Exploited"] = feedly_info["Exploited"]
        except Exception as e:
            print(f"[!] Error enriching CVE {cve_id} with Feedly: {e}")
            cve_entry["Public PoC"] = "Unknown"
            cve_entry["Exploited"] = "Unknown"

        # Optional: avoid hammering Feedly too fast
        time.sleep(1)
    print(f"[*] Feedly enrichment retrieved successfully!")

# Oracle specific functions
def start_oracle_workflow(patch_tuesday_date, month_selected):
    print(f"[*] Fetching Oracle Patch Tuesday data for {month_selected}...")
    oracle_cves = get_oracle_cves(month_selected)

    # MAY BREAK - if Feedly enrichment breaks and needs to be removed, remove the below 3 lines
    if oracle_cves:
        print(f"[*] Fetching Feedly enrichment... (this will take 1 second per CVE)")
        enrich_cves_with_feedly(oracle_cves)
        filename = f"Oracle-Patch-Tuesday-{month_selected}.xlsx"
        save_to_excel(oracle_cves, filename)

def get_oracle_cves(month_selected):
    print(f"[*] Fetching Oracle security page for {month_selected}...")

    # Convert 'Jul-2025' to 'jul2025' (Oracle's required format for the URL).
    full_month = datetime.strptime(month_selected, "%b-%Y").strftime("%b%Y").lower()

    # construct URL - adjust if you have a canonical oracle URL pattern
    oracle_url = ORACLE_SECURITY_NOTES_URL.format(full_month)

    try:
        response = make_vendor_request(oracle_url, full_month, "HTML")
    except Exception as e:
        # make_vendor_request already prints errors; ensure we gracefully exit
        print(f"[!] Could not fetch Oracle page for {month_selected}: {e}")
        return []

    # If make_vendor_request returned a Response-like object, extract text; if it returned json or str, adapt
    html = response.text if hasattr(response, "text") else str(response)

    # quick 404 check using response if available
    if hasattr(response, "status_code") and response.status_code == 404:
        print(f"[!] Oracle advisory page not found for {month_selected} (404). Skipping Oracle.")
        return []

    soup = BeautifulSoup(html, "html.parser")

    all_rows = []

    tables = soup.find_all("table")
    for table in tables:
        thead = table.find("thead")
        if not thead or len(thead.find_all("tr")) < 2:
            continue  # skip irrelevant tables

        # Extract headers
        headers = []
        top_row, sub_row = thead.find_all("tr")[:2]

        # Top-level headers (skip CVSS parent)
        for th in top_row.find_all("th"):
            if th.has_attr("colspan") and th.text.strip().startswith("CVSS"):
                # Second row: detailed CVSS headers
                for th in sub_row.find_all("th"):
                    headers.append(th.text.strip().replace("\n", " "))
                continue
            headers.append(th.text.strip().replace("\n", " "))

        # Normalize headers to lowercase for mapping
        headers_norm = [h.lower() for h in headers]

        # Only keep tables that contain CVE IDs and Base Score
        if not any("cve" in h for h in headers_norm) or not any("base score" in h for h in headers_norm):
            continue

        # Map columns
        col_map = {}
        for i, h in enumerate(headers_norm):
            if "cve" in h: col_map["cve"] = i
            if "product" in h: col_map["product"] = i
            if "component" in h: col_map["component"] = i
            if "supported versions" in h: col_map["versions"] = i
            if "base score" in h: col_map["cvss_score"] = i

        # Extract rows
        for tr in table.find("tbody").find_all("tr"):
            cells = tr.find_all(["th", "td"])
            def safe(idx):
                return cells[idx].get_text(" ", strip=True).replace("\xa0"," ") if idx is not None and idx < len(cells) else ""

            cve_text = safe(col_map.get("cve"))
            if not cve_text:
                found = re.findall(r"CVE-\d{4}-\d{4,7}", tr.get_text(" ", strip=True))
                if not found: 
                    continue
                cve_text = found[0]

            component = safe(col_map.get("component"))
            product = safe(col_map.get("product")) or component
            versions = safe(col_map.get("versions"))
            base_score = safe(col_map.get("cvss_score"))

            title = f"Component affected: {component}" if component else ""
            affected_products = f"{product} (versions: {versions})" if versions else product

            all_rows.append({
                "CVE": cve_text,
                "Title": title,
                "CVSS Score": base_score,
                "Public PoC": "Unknown",
                "Exploited": "Unknown",
                "Affected Products": affected_products,
                "Release Notes": oracle_url
            })

    if not all_rows:
        print(f"[!] No Oracle CVEs found for {month_selected}. Check if they have been released at {ORACLE_SECURITY_NOTES_URL.format(month_selected)}")
        return []
    
    print(f"[*] Oracle Patch Tuesday data for {month_selected} retrieved successfully")
    return all_rows


# SAP specific functions
# - get_sap_cves - nice and easy, we scrape a single page of all the information in the table
# - start_sap_workflow - start of SAP hell
def start_sap_workflow(patch_tuesday_date, month_selected):
    print(f"[*] Fetching SAP Patch Tuesday data for {month_selected}...")
    sap_cves = get_sap_cves(month_selected)

    if sap_cves:
        # MAY BREAK - if Feedly enrichment breaks and needs to be removed, remove the below 3 lines
        print(f"[*] Fetching Feedly enrichment... (this will take 1 second per CVE)")
        enrich_cves_with_feedly(sap_cves)

        filename = f"SAP-Patch-Tuesday-{month_selected}.xlsx"
        save_to_excel(sap_cves, filename)
    else:
        print(f"[!] No SAP CVEs found, no output produced")

def get_sap_cves(month_selected):
    sap_cves = []

    try:
        # Convert 'Jul-2025' to 'july-2025' (SAP's required format for the URL).
        full_month = datetime.strptime(month_selected, "%b-%Y").strftime("%B-%Y").lower()
        response = make_vendor_request(SAP_SECURITY_NOTES_URL.format(full_month), month_selected, "HTML")
        soup = BeautifulSoup(response.text, "html.parser")

        table = soup.find("table")
        if not table:
            print("[!] No SAP table found.")
            return []

        for row in table.find_all("tr")[1:]:  # Skip header row
            cols = row.find_all("td")
            if len(cols) < 4:
                continue

            note_link_tag = cols[0].find("a")
            release_note_url = note_link_tag["href"].strip() if note_link_tag and note_link_tag.has_attr("href") else ""

            # Title column: extract CVE and affected product
            title_col = cols[1]

            # Flatten all text and tags to check for non-CVE preamble
            first_significant_text = title_col.get_text(strip=True)
            first_cve_tag = title_col.find("a")

            if not first_cve_tag:
                print(f"[!] Skipping row with missing CVE hyperlink in SAP table at {SAP_SECURITY_NOTES_URL.format(full_month)}")
                continue

            # Only allow if the full title cell starts immediately with the CVE tag, skip any "Update to Security Note released on May 2025 Patch Day:"
            if not first_significant_text.startswith(f"[{first_cve_tag.get_text(strip=True)}]"):
                continue
            cve_id = first_cve_tag.get_text(strip=True).strip("[]") if first_cve_tag else "Unknown"
            title_text = title_col.get_text(" ", strip=True)

            # Extract product name
            product_match = re.search(r'Products?\s*.*?\s+(.*?)\s+Versions?',title_text,re.IGNORECASE | re.DOTALL)
            affected_product = product_match.group(1).strip() if product_match else "Unknown"
            if not product_match:
                print(f"[!] Could not extract product info from title: '{title_text[:80]}...' on {release_note_url}")
                affected_product = "Unknown"


            # Remove CVE and product/version lines for a cleaner title
            title_lines = title_text.split("Product –")[0].strip()
            clean_title = re.sub(r"\[\s*CVE-\d{4}-\d{4,7}\s*\]\s*", "", title_lines).strip()

            # CVSS score
            cvss_score = cols[3].get_text(strip=True)

            # Build final row for Excel
            sap_cves.append({
                "CVE": cve_id,
                "Title": clean_title,
                "CVSS Score": cvss_score,
                "Publicly Disclosed": "Unknown",
                "Exploited": "Unknown",
                "Affected Products": affected_product,
                "Publish Date": month_selected,
                "Release Notes": release_note_url
            })

    except Exception as e:
        print(f"[!] Error parsing SAP CVEs: {e}")

    print(f"[*] SAP Patch Tuesday data for {month_selected} retrieved successfully")
    return sap_cves

# Adobe specific functions
# Adobe sucks
# - get_adobe_product_links - starting from the initial security bulletin table page, scrape all the product security adversories for the month
# - extract_adobe_cves - navigate to each product security adversory page to get the CVE details
# - extract_cve_details_from_bulletin - once on a product page, extract the vulnerability table from the page and all the CVE details
# - resolve_column_indices - helper function bc Adobe is not consistent on its header titles in its vuln tables
# - find_vuln_details_table - helper function to find the vuln details table on the webpage
# - convert_adobe_rows - convert the info we have on the adoble vulns to standard format for output
# - start_adobe_workflow - start adoble hell

def get_adobe_product_links(patch_tuesday_date, month_selected):
    print(f"[*] Fetching Adobe Patch Tuesday data for {month_selected}...")
    response = make_vendor_request(ADOBE_SECURITY_BULLETIN_URL, month_selected, "HTML")
    print(f"[*] Adobe Patch Tuesday data for {month_selected} retrieved successfully")

    soup = BeautifulSoup(response.text, "html.parser")

    product_links = []
    # Parse each row in the security bulletin table
    table = soup.find("table")  # only the first table
    if not table:
        print("[!] No table found on bulletin page.")
        return []

    rows = table.find_all("tr")[1:]  # skip header
    for row in rows:
        cols = row.find_all("td")
        if len(cols) < 3:
            continue

        title_cell = cols[0]
        original_date_str = cols[1].get_text(strip=True)

        # Try parsing the date
        try:
            # Adobe is dumb and managed to output this abomination '08/012/2025' fix to '08/12/2025'
            raw_date = re.sub(r'\b0(\d{2})\b', r'\1', original_date_str)
            original_date = datetime.strptime(raw_date, "%m/%d/%Y").date()
        except ValueError:
            continue

        # Extract link and product name
        a_tag = title_cell.find("a", href=True)
        if not a_tag:
            continue

        href = a_tag["href"]
        product_url = href if href.startswith("http") else ADOBE_BASE_URL + href
        full_title = a_tag.get_text(strip=True)
        match = re.search(r"Security\s+update\s+available\s+for\s+(.+)$", full_title)
        product_name = match.group(1).strip() if match else full_title  # fallback
        if original_date == patch_tuesday_date:
            product_links.append({
                "Product": product_name,
                "URL": product_url,
                "Originally Posted": original_date_str
            })
    if not product_links:
        print(f"[!] No matching rows found in bulletin for selected month. Adobe cannot support historial data gathering. If you are trying to grab old data, this is why you are seeing this.")

    return product_links

def extract_adobe_cves(product_links):
    all_cves = []

    for entry in product_links:
        product = entry["Product"]
        url = entry["URL"]
        date = entry["Originally Posted"]
        print(f"[+] Extracting detailed CVEs for {product}")
        detailed_cves = extract_cve_details_from_bulletin(product, url, date)
        all_cves.extend(detailed_cves)
    return all_cves

def extract_cve_details_from_bulletin(product_name, bulletin_url, date):
    try:
        response = make_vendor_request(bulletin_url, product_name, "HTML")
        soup = BeautifulSoup(response.text, "html.parser")

        # Find the table by header
        table = find_vuln_details_table(soup)
        if not table:
            print(f"[!] Could not locate 'Vulnerability Details' table on {bulletin_url}")
            return []

        rows = table.find_all("tr")
        if not rows:
            print(f"[!] No rows found in vulnerability details table at {bulletin_url}")
            return []

        # Extract header cells (first row) and normalise whitespace/non-breaking spaces
        header_cells = rows[0].find_all(["td", "th"])
        headers = [cell.get_text(strip=True).replace("\xa0", " ") for cell in header_cells]

        # Expected column names (case-insensitive match)
        expected_cols = {
            "vuln_cat": ["Vulnerability Category"],
            "cvss_score": ["CVSS Base Score", "CVSS base score"],  # match both just in case
            "cve_number": ["CVE Numbers", "CVE Number", "CVE Number(s)"]
        }

        # Inline resolve_column_indices
        col_map = {}
        for key, possible_names in expected_cols.items():
            for name in possible_names:
                for idx, header in enumerate(headers):
                    if header.lower() == name.lower():
                        col_map[key] = idx
                        break
                if key in col_map:
                    break

        # Check if any expected column is missing
        missing_cols = [k for k in expected_cols if k not in col_map]
        if missing_cols:
            print(f"[!] Missing columns {missing_cols} in table headers at {bulletin_url}")
            print(f"    Found headers: {headers}")
            return []

        cve_details = []
        for row in rows[1:]:  # skip header row
            cells = row.find_all(["td", "th"])
            # Skip if row doesn't have enough columns
            if len(cells) <= max(col_map.values()):
                continue

            vuln_cat   = cells[col_map["vuln_cat"]].get_text(strip=True)
            cvss_score = cells[col_map["cvss_score"]].get_text(strip=True)
            cve_raw    = cells[col_map["cve_number"]].get_text(strip=True)

            # Extract CVE IDs from the raw text
            cve_list = re.findall(r"CVE-\d{4}-\d{4,7}", cve_raw, re.IGNORECASE)

            for cve in cve_list:
                cve_details.append({
                    "CVE": cve,
                    "Title": vuln_cat,
                    "CVSS Score": cvss_score,
                    "Publicly Disclosed": "Unknown",
                    "Exploited": "Unknown",
                    "Affected Products": product_name,
                    "Publish Date": date,
                    "Release Notes": bulletin_url
                })

        return cve_details

    except Exception as e:
        print(f"[!] Failed to extract detailed CVE info from {bulletin_url}: {e}")
        return []
    
def find_vuln_details_table(soup):
    # Look for all <h2> elements
    for h2 in soup.find_all("h2"):
        if "vulnerability details" in h2.get_text(strip=True).lower():
            # Search forward in the document for the first <table>
            next_el = h2
            while next_el:
                next_el = next_el.find_next()
                if next_el.name == "table":
                    return next_el
    return None

def start_adobe_workflow(patch_tuesday_date, month_selected):
    product_links = get_adobe_product_links(patch_tuesday_date, month_selected)
    all_cves = extract_adobe_cves(product_links)

    if all_cves:
        # MAY BREAK - if Feedly enrichment breaks and needs to be removed, remove the below 3 lines
        print(f"[*] Fetching Feedly enrichment... (this will take 1 second per CVE)")
        enrich_cves_with_feedly(all_cves)

        filename = f"Adobe-Patch-Tuesday-{month_selected}.xlsx"
        save_to_excel(all_cves, filename)
    else:
        print(f"[!] No Adobe CVEs found, no output produced")

# Microsoft specific functions
#  - resolve_product_name_microsoft - the affected products are just listed as IDs in the CVE information returned by the API, gotta convert to product name by checking a different section of the API response
# - extract_vulnerability_info_microsoft - get all the juicy information in a microsoft specific way
# - start_microsoft_workflow - start of microsoft hell

def resolve_product_name_microsoft(product_id, doc):
    for product in doc.get("ProductTree", {}).get("FullProductName", []):
        if product["ProductID"] == product_id:
            return product["Value"]
    return product_id

def extract_vulnerability_info_microsoft(doc, patch_tuesday_date):
    rows = []
    start_of_month = patch_tuesday_date.replace(day=1)

    for vuln in doc.get("Vulnerability", []):
        # check RevisionHistory[0]["Date"] to get the date published 
        rev_history = vuln.get("RevisionHistory", [])
        if not rev_history:
            continue
        try:
            pub_date_str = rev_history[0]["Date"]
            pub_date = datetime.fromisoformat(pub_date_str.replace("Z", "+00:00")).date()
        except Exception:
            continue  # skip if date is malformed

        if not (pub_date == patch_tuesday_date): # MAY BREAK - if you need all vulns in a month with no end date, replace line with if not (start_of_month <= pub_date)
            continue  # skip CVEs not published from the start of the month to patch tuesday aka not included in the security update notes. We specify an end date in case the script is being run later in the month

        cve = vuln.get("CVE", "N/A")
        disclosed = "No"
        exploited = "No"

        for threat in vuln.get("Threats", []):
            desc = threat.get("Description", {}).get("Value", "")
            if "Publicly Disclosed:" in desc and "Exploited:" in desc:
                # Example: "Publicly Disclosed:No;Exploited:Yes;..."
                parts = dict(item.split(":") for item in desc.split(";") if ":" in item)
                disclosed = parts.get("Publicly Disclosed", "No")
                exploited = parts.get("Exploited", "No")
                break 


        cvss_base = "N/A"
        for score in vuln.get("CVSSScoreSets", []):
            if "BaseScore" in score:
                cvss_base = score["BaseScore"]
                break

        title = vuln.get("Title", {}).get("Value", "N/A")

        product_names = set()
        for product_status in vuln.get("ProductStatuses", []):
            for pid in product_status.get("ProductID", []):
                product_name = resolve_product_name_microsoft(pid, doc)
                if product_name == pid:
                    print(f"[!] Product ID '{pid}' could not be resolved for CVE {cve}")
                product_names.add(product_name)

        release_note = MSRC_API_CVE_URL.format(cve)

        rows.append({
            "CVE": cve,
            "Title": title,
            "CVSS Score": cvss_base,
            "Public PoC": disclosed,
            "Exploited": exploited,
            "Affected Products": ", ".join(sorted(product_names)),
            "Publish Date": pub_date,
            "Release Notes": release_note #RELEASE NOTE MUST ALWAYS BE LAST COLUMN
        })

    return rows

def start_microsoft_workflow(year, month_abbr, patch_tuesday_date):
    # Normalize to Microsoft's expected ID format: 2025-Jul
    month_normalized = f"{year}-{month_abbr}"
    
    print(f"[*] Fetching Microsoft Patch Tuesday data for {month_normalized}...")
    doc = make_vendor_request(MSRC_API_CVRF_URL.format(month_normalized), month_normalized, "JSON")
    print(f"[*] Microsoft Patch Tuesday data for {month_normalized} retrieved successfully")

    print(f"[*] Sorting Microsoft API data for {month_normalized}...")
    rows = extract_vulnerability_info_microsoft(doc,patch_tuesday_date)
    print(f"[*] Microsoft API data for {month_normalized} sorted successfully")

    filename = f"Microsoft_Patch_Tuesday_{month_abbr}-{year}.xlsx"
    save_to_excel(rows, filename)

# main is main

def main():
    parser = argparse.ArgumentParser(description="Fetch Patch Tuesday vulnerability data and export to Excel.")
    parser.add_argument("--microsoft", action="store_true", help="Fetch data from Microsoft CVRF API feed")
    parser.add_argument("--adobe", action="store_true", help="Fetch data from Adobe Security Bulletin and enrich with Feedly")
    parser.add_argument("--sap", action="store_true", help="Fetch data from SAP Security Notes and enrich with Feedly")
    parser.add_argument("--oracle", action="store_true", help="Fetch data from Oracle Security Notes and enrich with Feedly")
    parser.add_argument("--all", action="store_true", help="Get all data from Microsoft, Adobe, SAP and Oracle")
    parser.add_argument("month", help="Target month in format e.g. Jul-2025")

    args = parser.parse_args()

    if not args.month:
        print("Error: You must provide a month in the format 'Jul-2025'.")
        sys.exit(1)
    
    month_selected = args.month
    month_abbr, year = month_selected.split("-")
    patch_tuesday_date = get_patch_tuesday(int(year), datetime.strptime(month_abbr, "%b").month)

    # If --all is set, run everything
    if args.all:
        start_microsoft_workflow(year, month_abbr, patch_tuesday_date)
        start_adobe_workflow(patch_tuesday_date, month_selected)
        start_sap_workflow(patch_tuesday_date, month_selected)
        start_oracle_workflow(patch_tuesday_date, month_selected)
    # If at least one individual flag is set
    elif args.microsoft or args.adobe or args.sap or args.oracle:
        if args.microsoft:
            start_microsoft_workflow(year, month_abbr, patch_tuesday_date)
        if args.adobe:
            start_adobe_workflow(patch_tuesday_date, month_selected)
        if args.sap:
            start_sap_workflow(patch_tuesday_date, month_selected)
        if args.oracle:
            start_oracle_workflow(patch_tuesday_date, month_selected)
    # No valid vendor flag passed
    else:
        print("Error: Please specify a vendor. Supported options: --microsoft  --adobe  --sap  --all")
        sys.exit(1)

if __name__ == "__main__":
    main()
